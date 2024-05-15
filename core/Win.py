import os

import numpy as np
import openpyxl as px
from openpyxl.styles import Alignment
import pyqtgraph as pg
from PyQt5.QtGui import QIntValidator, QFont, QColor
from PyQt5.QtCore import Qt, QPoint, QCoreApplication
from PyQt5.QtWidgets import QApplication, QMessageBox, QFileDialog, QDialog, QTableWidgetItem, QMainWindow, QMenu, QTableWidgetItem, QHeaderView

from utils.get_turning_point import get_turning_point
from utils.data_expansion import data_expansion
from utils.data_monotonisation import data_monotonisation
from utils.smooth import smooth
from utils.loops_division import loops_division
from utils.get_dissipated_energy import get_dissipated_energy
from utils.get_skeleton import get_skeleton
from utils.get_residual_deformation import get_residual_deformation
from utils.get_skeleton_curve import get_skeleton_curve_1, get_skeleton_curve_2
from ui.MainWin import Ui_MainWindow
from ui.WinData import Ui_WinData
from ui.WinAbout import Ui_WinAbout
from ui.WinHelp import Ui_WinHelp


class MainWin(QMainWindow):
    Version = 'V3.0'
    date = '2024.5.15'
    u1, u2, u3, u4, u5, u6, u7, u7_1 = None, None, None, None, None, None, None, None
    F1, F2, F3, F4, F5, F6, F7, u7_1 = None, None, None, None, None, None, None, None
    ok1, ok2, ok3, ok4, ok5, ok6, ok7, ok7_1 = False, False, False, False, False, False, False, False
    u_import, F_import = False, False  # 是否存在位移或力数据
    pg_antialias = False
    
    def __init__(self):
        super().__init__()
        self.init_pg()
        self.init_ui()

    def init_pg(self):
        """初始化画图设置"""
        self.pen1 = pg.mkPen(color=(68, 114, 196), width=2)  # matplotlib蓝色:30, 120, 180
        self.pen2 = pg.mkPen(color='orange', width=2)  # 橙色
        self.pen3 = pg.mkPen(color='green', width=2)  # 绿色
        self.pen4 = pg.mkPen(color=(211, 211, 211), width=2)  # 亮灰色
        self.pen5 = pg.mkPen(color='red', width=2)  # 红色
        self.pen_axis = pg.mkPen(color='black', width=1)  # 坐标轴
        self.pen_grid = pg.mkPen(color='grey', width=1)
        self.font_axis = QFont('Arial')
        self.font_axis.setPixelSize(15)  # 坐标轴字体大小

    def init_ui(self):
        self.ui = Ui_MainWindow()  # 主窗口
        self.ui.setupUi(self)
        self.setWindowTitle(f'滞回曲线处理软件 {MainWin.Version}')
        # tab 1
        self.ui.radioButton.pressed.connect(lambda: self.ui.pushButton.setEnabled(True))
        self.ui.radioButton.pressed.connect(lambda: self.ui.pushButton_2.setEnabled(True))
        self.ui.radioButton.pressed.connect(lambda: self.ui.pushButton_3.setEnabled(False))
        self.ui.radioButton_2.pressed.connect(lambda: self.ui.pushButton.setEnabled(False))
        self.ui.radioButton_2.pressed.connect(lambda: self.ui.pushButton_2.setEnabled(False))
        self.ui.radioButton_2.pressed.connect(lambda: self.ui.pushButton_3.setEnabled(True))
        self.ui.checkBox_6.clicked.connect(self.reverse_disp)
        self.ui.pushButton_8.clicked.connect(self.clear_data)
        self.ui.pushButton.clicked.connect(self.get_disp)
        self.ui.pushButton_2.clicked.connect(self.get_force)
        self.ui.pushButton_3.clicked.connect(self.get_hysteretic)
        self.ui.pushButton_4.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
        self.ui.pushButton_4.clicked.connect(self.tab1_finished)
        self.pg1 = self.replace_to_pyqtgraph(self.ui.graphicsView, self.ui.verticalLayout_23, 1)
        self.pg2 = self.replace_to_pyqtgraph(self.ui.graphicsView_2, self.ui.verticalLayout_23, 3)
        for i in range(4):
            self.ui.verticalLayout_23.setStretch(i, [0, 15, 0, 7][i])
        # tab 2
        self.ui.pushButton_5.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(0))
        self.ui.pushButton_6.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
        self.ui.pushButton_6.clicked.connect(self.tab2_finished)
        self.ui.checkBox_4.toggled.connect(self.delete_duplicate_point)
        self.pg3 = self.replace_to_pyqtgraph(self.ui.graphicsView_3, self.ui.verticalLayout_5, 1)
        self.pg4 = self.replace_to_pyqtgraph(self.ui.graphicsView_4, self.ui.verticalLayout_5, 3)
        for i in range(4):
            self.ui.verticalLayout_5.setStretch(i, [0, 1, 0, 1][i])
        # tab 3
        self.ui.pushButton_9.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(1))
        self.ui.pushButton_10.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
        self.ui.pushButton_10.clicked.connect(self.tab3_finished)
        self.ui.horizontalSlider.valueChanged.connect(self.change_slider_1)
        self.ui.horizontalSlider_2.valueChanged.connect(self.change_slider_2)
        self.ui.lineEdit_9.editingFinished.connect(self.set_slider_limit_1)
        self.ui.lineEdit_8.editingFinished.connect(self.set_slider_limit_2)
        self.ui.lineEdit_10.editingFinished.connect(self.set_slider_1)
        self.ui.lineEdit_11.editingFinished.connect(self.set_slider_2)
        self.ui.lineEdit_9.setValidator(QIntValidator())
        self.ui.lineEdit_10.setValidator(QIntValidator())
        self.ui.lineEdit_11.setValidator(QIntValidator())
        self.pg5 = self.replace_to_pyqtgraph(self.ui.graphicsView_5, self.ui.verticalLayout_6, 1)
        self.ui.checkBox_5.stateChanged.connect(self.checkBox_5_changed)
        self.ui.pushButton_33.clicked.connect(self.delete_clicked_points)
        self.ui.pushButton_34.clicked.connect(self.reselect_point)
        # tab 4
        self.ui.pushButton_11.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(2))
        self.ui.pushButton_12.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
        self.ui.pushButton_12.clicked.connect(self.tab4_finished)
        self.ui.checkBox.stateChanged.connect(self.checkBox_state_changed)
        self.ui.lineEdit_12.editingFinished.connect(self.data_expansion)
        # tab 5
        self.ui.pushButton_13.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(3))
        self.ui.pushButton_14.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(5))
        self.ui.pushButton_14.clicked.connect(self.tab5_finished)
        self.ui.checkBox_2.stateChanged.connect(self.checkBox2_state_changed)
        self.ui.lineEdit_13.editingFinished.connect(self.curve_smooth)
        self.pg6 = self.replace_to_pyqtgraph(self.ui.graphicsView_6, self.ui.verticalLayout_12, 1)
        self.ui.lineEdit_13.setValidator(QIntValidator())
        for i in range(3):
            self.ui.verticalLayout_6.setStretch(i, [0, 0, 1][i])
        # tab 6
        self.ui.pushButton_15.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(4))
        self.ui.pushButton_16.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(6))
        self.ui.pushButton_16.clicked.connect(self.tab6_finished)
        self.pg7 = self.replace_to_pyqtgraph(self.ui.graphicsView_7, self.ui.verticalLayout_17, 1)
        self.pg8 = self.replace_to_pyqtgraph(self.ui.graphicsView_8, self.ui.verticalLayout_18, 1)
        self.pg9 = self.replace_to_pyqtgraph(self.ui.graphicsView_9, self.ui.verticalLayout_19, 1)
        self.pg10 = self.replace_to_pyqtgraph(self.ui.graphicsView_10, self.ui.verticalLayout_20, 1)
        self.ui.radioButton_3.toggled.connect(self.get_loops_and_energy)
        self.ui.checkBox_3.toggled.connect(self.get_loops_and_energy)
        self.ui.radioButton_5.clicked.connect(self.get_loops_and_energy)
        self.ui.radioButton_6.clicked.connect(self.get_loops_and_energy)
        self.ui.radioButton_7.clicked.connect(self.get_loops_and_energy)
        self.ui.radioButton_10.toggled.connect(self.get_loops_and_energy)
        # tab 7
        self.ui.pushButton_25.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(5))
        self.ui.pushButton_26.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(7))
        self.ui.pushButton_26.clicked.connect(self.tab7_finished)
        self.pg12 = self.replace_to_pyqtgraph(self.ui.graphicsView_12, self.ui.verticalLayout_27, 1)
        self.ui.comboBox.currentIndexChanged.connect(self.comboBox_changed)
        self.ui.pushButton_7.clicked.connect(self.previous_loop)
        self.ui.pushButton_27.clicked.connect(self.next_loop)
        self.ui.pushButton_28.clicked.connect(self.plot_all_loop)
        # tab 7_1
        self.ui.pushButton_29.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(6))
        self.ui.pushButton_30.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(8))
        self.ui.pushButton_30.clicked.connect(self.tab7_1_finished)
        self.pg13 = self.replace_to_pyqtgraph(self.ui.graphicsView_13, self.ui.verticalLayout_29, 1)
        self.pg14 = self.replace_to_pyqtgraph(self.ui.graphicsView_14, self.ui.verticalLayout_30, 1)
        self.pg15 = self.replace_to_pyqtgraph(self.ui.graphicsView_15, self.ui.verticalLayout_31, 1)
        self.ui.radioButton_8.clicked.connect(lambda: self.ui.comboBox_5.setEnabled(True))
        self.ui.radioButton_8.clicked.connect(lambda: self.ui.lineEdit_27.setEnabled(False))
        self.ui.radioButton_9.clicked.connect(lambda: self.ui.comboBox_5.setEnabled(False))
        self.ui.radioButton_9.clicked.connect(lambda: self.ui.lineEdit_27.setEnabled(True))
        self.ui.pushButton_31.clicked.connect(self.clicked_add_to_scheme)
        self.scheme = None
        self.is_clear = False
        self.ui.pushButton_32.clicked.connect(self.clear_all_scheme)
        self.ui.comboBox_2.currentIndexChanged.connect(self.plot_skeleton)
        self.ui.comboBox_3.currentIndexChanged.connect(self.plot_stiffness_degradation)
        self.ui.comboBox_4.currentIndexChanged.connect(self.plot_strength_degradation)
        self.data_skeleton_curve = []  # [[[u], [F]], [[u], [F]], ...]
        self.data_stiffness_degradation = []
        self.data_strength_degradation = []
        self.data_scheme = []
        # tab 8
        self.ui.pushButton_23.clicked.connect(lambda: self.ui.tabWidget.setCurrentIndex(7))
        self.ui.pushButton_17.clicked.connect(self.data_hysteretic_curve)
        self.ui.pushButton_18.clicked.connect(self.data_skeleton_points)
        self.ui.pushButton_19.clicked.connect(self.data_Es)
        self.ui.pushButton_20.clicked.connect(self.data_Ea)
        self.ui.pushButton_21.clicked.connect(self.data_zeta)
        self.ui.pushButton_22.clicked.connect(self.data_residual_deformation)
        self.ui.pushButton_24.clicked.connect(self.export_all_data)
        self.pg11 = self.replace_to_pyqtgraph(self.ui.graphicsView_11, self.ui.verticalLayout_16, 2)
        # Menu and statusbar
        self.current_tab_index = 0  # 当前tab的索引号
        self.ui.tabWidget.tabBarClicked.connect(self.tab_clicked)
        self.ui.statusbar.showMessage(f'滞回曲线处理软件 {MainWin.Version}')
        self.ui.action_2.triggered.connect(self.show_WinHelp)
        self.ui.action_3.triggered.connect(self.show_WinAbout)

    def tab_clicked(self, index: int):
        """当tab被点击时"""
        print(f'tab clicked (index = {index})')
        if index == 1:
            self.tab1_finished()
        if index == 2:
            self.tab2_finished()
        if index == 3:
            self.tab3_finished()
        if index == 4:
            self.tab4_finished()
        if index == 5:
            self.tab5_finished()
        if index == 6:
            self.tab6_finished()
        if index == 7:
            self.tab7_finished()
        if index == 8:
            self.tab7_1_finished()

    def set_current_tab_index(self):
        """设置当前的tab索引号"""
        self.current_tab_index = self.ui.tabWidget.currentIndex()

    def replace_to_pyqtgraph(self, graphicsView, layout, index):
        """将graphicsView控件替换为pyqtgrapg"""
        layout.removeWidget(graphicsView)
        graphicsView.deleteLater()
        pg_widget = pg.PlotWidget()
        pg_widget.setBackground('w')
        # 显示上轴和右轴
        pg_widget.showAxis('top')
        pg_widget.showAxis('right')
        # 设置坐标轴宽度
        pg_widget.getAxis('bottom').setPen(self.pen_axis)
        pg_widget.getAxis('top').setPen(self.pen_axis)
        pg_widget.getAxis('left').setPen(self.pen_axis)
        pg_widget.getAxis('right').setPen(self.pen_axis)
        # 更改坐标轴刻度字体大小
        pg_widget.getAxis('bottom').setTickFont(self.font_axis)
        pg_widget.getAxis('top').setTickFont(self.font_axis)
        pg_widget.getAxis('left').setTickFont(self.font_axis)
        pg_widget.getAxis('right').setTickFont(self.font_axis)
        # 更改坐标轴字体颜色
        pg_widget.getAxis('bottom').setTextPen(QColor('black'))
        pg_widget.getAxis('top').setTextPen(QColor('black'))
        pg_widget.getAxis('left').setTextPen(QColor('black'))
        pg_widget.getAxis('right').setTextPen(QColor('black'))
        # 不显示上轴和右轴的刻度标签
        pg_widget.getAxis('top').setTicks([])
        pg_widget.getAxis('right').setTicks([])
        # 显示网格线
        pg_widget.showGrid(True, True)
        # 添加到原layout中
        layout.insertWidget(index, pg_widget)
        return pg_widget

    @staticmethod
    def MyDel(*args):
        for arg in args:
            try:
                exec(f'del {arg}')
            except:
                pass

    # --------------------------------------------------- tab 1 ---------------------------------------------------

    def get_disp(self):
        """点击 - 导入位移"""
        u_file = QFileDialog.getOpenFileName(self, '导入位移', '', '文本文档 (*.txt *.out)')[0]
        u_file_dir = u_file.split('/')[-1]
        if not u_file:
            return 0
        try:
            u1_temp = np.loadtxt(u_file)
        except:
            QMessageBox.warning(self, '警告', '无法导入数据，请检查数据格式！')
            return 0
        if len(u1_temp) == 0:
            QMessageBox.warning(self, '警告', '数据为空！')
            return 0
        try:
            self.plot_protocal(self.pg2, np.arange(0, len(u1_temp), 1), u1_temp)
        except:
            QMessageBox.warning(self, '警告', '无法绘制加载制度！')
            return 0
        self.ui.label_13.setText(f'位移数据：{u_file_dir}')
        MainWin.u1 = u1_temp
        MainWin.u_import = True
        if not self.MainWin_plot():
            return 0
        self.data_statistics()

    def get_force(self):
        """点击 - 导入力"""
        F_file = QFileDialog.getOpenFileName(self, '导入力', '', '文本文档 (*.txt *.out)')[0]
        F_file_dir = F_file.split('/')[-1]
        if not F_file:
            return 0
        try:
            F1_temp = np.loadtxt(F_file)
        except:
            QMessageBox.warning(self, '警告', '无法导入数据，请检查数据格式！')
            return 0
        if len(F1_temp) == 0:
            QMessageBox.warning(self, '警告', '没有数据！')
            return 0
        MainWin.F1 = F1_temp
        MainWin.F_import = True
        self.ui.label_14.setText(f'力数据：{F_file_dir}')
        if not self.MainWin_plot():
            return 0
        self.data_statistics()

    def get_hysteretic(self):
        """点击 - 导入滞回曲线"""
        uF_file = QFileDialog.getOpenFileName(self, '导入滞回曲线', '', '文本文档 (*.txt *.out)')[0]
        uF_file_dir = uF_file.split('/')[-1]
        if not uF_file:
            return 0
        try:
            data = np.loadtxt(uF_file)
            u1_temp = data[:, 0]
            F1_temp = data[:, 1]
        except:
            QMessageBox.warning(self, '警告', '无法导入数据，请检查数据格式！')
            return 0
        if len(u1_temp) == 0 or len(F1_temp) == 0:
            QMessageBox.warning(self, '警告', '数据为空！')
            return 0
        try:
            self.plot_protocal(self.pg2, np.arange(0, len(u1_temp), 1), u1_temp)
        except:
            QMessageBox.warning(self, '警告', '无法绘制加载制度！')
            return 0
        self.ui.label_13.setText(f'位移数据：{uF_file_dir}')
        self.ui.label_14.setText(f'力数据：{uF_file_dir}')
        MainWin.u1, MainWin.F1 = u1_temp, F1_temp
        MainWin.u_import, MainWin.F_import = True, True
        if not self.MainWin_plot():
            return 0
        self.data_statistics()

    def plot_protocal(self, pg_i, x, y):
        """绘制加载制度"""
        pg_i.clear()
        pg_i.plot(x, y, pen=self.pen1, antialias=MainWin.pg_antialias)
        pg_i.autoRange()

    def MainWin_plot(self):
        """绘制曲线，初步统计"""
        if all([MainWin.u_import, MainWin.F_import]):
            try:
                self.pg1.clear()
                self.pg1.plot(MainWin.u1, MainWin.F1, pen=self.pen1, antialias=MainWin.pg_antialias)
                self.pg1.autoRange()
                MainWin.ok1 = True
                self.ui.checkBox_6.setEnabled(True)
            except:
                QMessageBox.warning(self, '警告', '无法绘制滞回曲线！')
                return 0
        return 1
    
    def reverse_disp(self):
        if MainWin.u_import and MainWin.F_import:
            MainWin.u1 = -MainWin.u1
            MainWin.F1 = -MainWin.F1
            self.plot_protocal(self.pg2, np.arange(0, len(MainWin.u1), 1), MainWin.u1)
            self.MainWin_plot()

    def clear_data(self):
        """清除所有数据"""
        MainWin.u1, MainWin.F1 = None, None
        MainWin.u_import, MainWin.F_import = None, None
        MainWin.ok1 = False
        self.pg1.clear()
        self.pg2.clear()
        self.ui.label_13.setText('位移数据：')
        self.ui.label_14.setText('力数据：')
        self.ui.lineEdit.clear()
        self.ui.lineEdit_2.clear()
        self.ui.lineEdit_3.clear()
        self.ui.lineEdit_4.clear()
        self.ui.lineEdit_5.clear()
        self.ui.checkBox_6.setChecked(False)
        self.ui.checkBox_6.setEnabled(False)
        self.ui.radioButton_2.setChecked(True)
        self.ui.pushButton_3.setEnabled(True)
        self.ui.pushButton.setEnabled(False)
        self.ui.pushButton_2.setEnabled(False)
        self.clear_tab2()
        self.clear_tab3()
        self.clear_tab4()
        self.clear_tab5()
        self.clear_tab6()
        self.clear_tab7()
        self.clear_tab7_1()
        self.clear_tab8()

    def data_statistics(self):
        if MainWin.u_import:
            self.ui.lineEdit.setText(str(max(MainWin.u1)))
            self.ui.lineEdit_2.setText(str(min(MainWin.u1)))
        if MainWin.F_import:
            self.ui.lineEdit_3.setText(str(max(MainWin.F1)))
            self.ui.lineEdit_4.setText(str(min(MainWin.F1)))
        if all([MainWin.u_import, MainWin.F_import]):
            self.ui.lineEdit_5.setText(str(len(MainWin.u1)))

    def tab1_finished(self):
        self.tab2_start()

    # --------------------------------------------------- tab 2 ---------------------------------------------------

    def tab2_start(self):
        self.plot2_old()
        self.delete_duplicate_point()

    def plot2_old(self):
        """绘制旧加载制度"""
        if MainWin.ok1:
            self.pg3.clear()
            self.pg3.plot(np.arange(0, len(MainWin.u1), 1), MainWin.u1, pen=self.pen1)
            self.pg3.autoRange()
            self.ui.lineEdit_6.setText(str(len(MainWin.u1)))
            self.pg4.clear()

    def delete_duplicate_point(self):
        """删除重复点"""
        if MainWin.ok1:
            self.u2_temp, self.F2_temp = np.array([MainWin.u1[0]]), np.array([MainWin.F1[0]])
            if not (self.u2_temp[0] == 0 and self.F2_temp[0] == 0):
                if self.ui.checkBox_4.isChecked():
                    self.u2_temp, self.F2_temp = np.insert(self.u2_temp, 0, 0),np.insert(self.F2_temp, 0, 0)
            for i in range(1, len(MainWin.u1)):
                if MainWin.u1[i] != self.u2_temp[-1]:
                    self.u2_temp = np.append(self.u2_temp, MainWin.u1[i])
                    self.F2_temp = np.append(self.F2_temp, MainWin.F1[i])
            self.pg4.clear()
            self.pg4.plot(np.arange(0, len(self.u2_temp), 1), self.u2_temp, pen=self.pen1)
            self.pg4.autoRange()
            self.ui.lineEdit_7.setText(str(len(self.u2_temp)))
    
    def tab2_finished(self):
        if MainWin.ok1:
            MainWin.u2, MainWin.F2 = self.u2_temp, self.F2_temp
            MainWin.ok2 = True
            self.tab3_start()

    def clear_tab2(self):
        MainWin.u2, MainWin.F2 = None, None
        MainWin.ok2 = False
        self.ui.checkBox_4.setChecked(True)
        self.ui.lineEdit_6.clear()
        self.ui.lineEdit_7.clear()
        self.pg3.clear()
        self.pg4.clear()
        self.MyDel('self.u2_temp', 'self.F2_temp')

    # --------------------------------------------------- tab 3 ---------------------------------------------------

    def tab3_start(self):
        self.plot_turning_points()

    def checkBox_5_changed(self):
        if self.ui.checkBox_5.isChecked():
            self.ui.pushButton_10.setEnabled(True)
        else:
            self.ui.pushButton_10.setEnabled(False)
        
    def change_slider_1(self):
        slider_value1 = self.ui.horizontalSlider.value()
        self.ui.lineEdit_10.setText(str(slider_value1))
        self.plot_turning_points()

    def change_slider_2(self):
        slider_value2 = self.ui.horizontalSlider_2.value()
        self.ui.lineEdit_11.setText(str(slider_value2))
        self.plot_turning_points()

    def set_slider_limit_1(self):
        """设置起始宽度滑条上限"""
        limit1 = self.ui.lineEdit_9.text()
        if int(limit1) < 2:
            self.ui.lineEdit_9.setText('2')
            limit1 = '2'
        self.ui.label_19.setText(limit1)
        self.ui.horizontalSlider.setMaximum(int(limit1))
        self.ui.horizontalSlider.setTickInterval(int(int(limit1) / 10))
        self.ui.lineEdit_10.setText(str(self.ui.horizontalSlider.value()))

    def set_slider_limit_2(self):
        """设置终止宽度滑条上限"""
        limit2 = self.ui.lineEdit_8.text()
        if int(limit2) < 3:
            self.ui.lineEdit_8.setText('2')
            limit2 = '2'
        self.ui.label_20.setText(limit2)
        self.ui.horizontalSlider_2.setMaximum(int(limit2))
        self.ui.horizontalSlider_2.setTickInterval(int(int(limit2) / 10))
        self.ui.lineEdit_11.setText(str(self.ui.horizontalSlider_2.value()))

    def set_slider_1(self):
        win1 = int(self.ui.lineEdit_10.text())
        win1 = max(2, win1)
        self.ui.lineEdit_10.setText(str(win1))
        self.ui.horizontalSlider.setValue(win1)

    def set_slider_2(self):
        win2 = int(self.ui.lineEdit_11.text())
        win2 = max(2, win2)
        self.ui.lineEdit_11.setText(str(win2))
        self.ui.horizontalSlider_2.setValue(win2)

    def plot_turning_points(self, delete_points=None):
        if MainWin.ok2:
            _, self.tag = get_turning_point(MainWin.u2, self.ui.horizontalSlider.value(), self.ui.horizontalSlider_2.value())
            if delete_points:
                # 如果有删除点
                self.current_tag = np.delete(self.tag, delete_points)
                self.current_x = np.delete(np.arange(0, len(MainWin.u2), 1)[self.tag], delete_points)  # 删除点后的反向点
                self.current_y = np.delete(MainWin.u2[self.tag], delete_points)
            else:
                self.idx_point_selected = []  # 所选点的索引
                self.x_point_selected, self.y_point_selected = [], []  # 所选点的坐标
                self.current_tag = self.tag
                self.current_x = np.arange(0, len(MainWin.u2), 1)[self.current_tag]
                self.current_y = MainWin.u2[self.tag]
            self.pg5.clear()
            self.line1 = pg.PlotCurveItem(np.arange(0, len(MainWin.u2), 1), MainWin.u2,  pen=self.pen4)  # 位移序列
            self.pg5.addItem(self.line1)
            self.line2 = pg.ScatterPlotItem(size=8, pen=None, brush='red')  # 反向点
            self.line2.addPoints(self.current_x, self.current_y)
            self.pg5.addItem(self.line2)
            pen = pg.mkPen('orange')
            pen.setWidth(4)
            self.line3 = pg.ScatterPlotItem(size=17, pen=pen, brush=None)  # 删除点
            if delete_points:
                self.line3.addPoints(self.x_point_selected, self.y_point_selected)
            self.pg5.addItem(self.line3)
            self.line2.sigClicked.connect(self.point_clicked)

    def point_clicked(self, plot, points):
        for point in points:
            pos = point.pos()
            x, y = pos.x(), pos.y()
            index = np.where((np.arange(0, len(MainWin.u2), 1)[self.tag] == x) & (MainWin.u2[self.tag] == y))[0]
            if (len(index) > 0) and (index not in self.idx_point_selected):
                self.idx_point_selected.append(index)
                self.x_point_selected.append(x)
                self.y_point_selected.append(y)
                self.line3.addPoints(self.x_point_selected, self.y_point_selected)

    def delete_clicked_points(self):
        self.plot_turning_points(delete_points=self.idx_point_selected)
        
    def reselect_point(self):
        self.plot_turning_points()

    def tab3_finished(self):
        if MainWin.ok2:
            MainWin.u3, MainWin.F3 = MainWin.u2, MainWin.F2
            MainWin.ok3 = True
            self.tab4_start()
    
    def clear_tab3(self):
        MainWin.u3, MainWin.F3 = None, None
        MainWin.ok3 = False
        self.ui.lineEdit_9.setText('200')
        self.ui.lineEdit_8.setText('200')
        self.ui.lineEdit_10.setText('100')
        self.ui.lineEdit_11.setText('100')
        self.ui.horizontalSlider.setMaximum(200)
        self.ui.horizontalSlider_2.setMaximum(200)
        self.ui.horizontalSlider.setValue(100)
        self.ui.horizontalSlider_2.setValue(100)
        # self.ui.horizontalSlider.set
        self.pg5.clear()
        self.ui.checkBox_5.setChecked(False)
        self.MyDel('self.tag', 'self.current_tag', 'self.current_x', 'self.current_y',\
                   'self.idx_point_selected', 'self.x_point_selected', 'self.y_point_selected',\
                    'self.line1', 'self.line2', 'self.line3')

    # --------------------------------------------------- tab 4 ---------------------------------------------------

    def tab4_start(self):
        self.tag = self.current_tag
        self.get_len_before_process()

    def checkBox_state_changed(self, state):
        if state == Qt.Checked:
            self.ui.label_26.setEnabled(True)
            self.ui.lineEdit_12.setEnabled(True)
            self.data_expansion()
            try:
                self.ui.label_28.setText(f'扩充后数据长度：{str(len(self.u4_temp_expan))}')
            except:
                self.ui.label_28.setText('扩充后数据长度：')
        else:
            self.ui.label_26.setEnabled(False)
            self.ui.lineEdit_12.setEnabled(False)
            self.ui.label_28.setText('扩充后数据长度：不扩充')

    def get_len_before_process(self):
        # 获取数据处理前长度
        if MainWin.ok3:
            self.ui.label_27.setText(f'处理前数据长度：{str(len(MainWin.u3))}')
            self.data_monotonisation()

    def data_monotonisation(self):
        # 数据单调化
        if MainWin.ok3:
            self.u4_temp, self.F4_temp = data_monotonisation(MainWin.u3, MainWin.F3, self.tag)
            self.ui.label_29.setText(f'单调化后数据长度：{str(len(self.u4_temp))}')

    def data_expansion(self):
        # 进行数据扩充
        if MainWin.ok3:
            min_du_str = self.ui.lineEdit_12.text()
            try:
                self.min_du = float(min_du_str)
                if not self.min_du > 0:
                    self.ui.label_28.setText('扩充后数据长度：')
                    return 0
            except:
                self.ui.label_28.setText('扩充后数据长度：')
                return 0
            self.u4_temp_expan, self.F4_temp_expan = data_expansion(self.u4_temp, self.F4_temp, self.min_du)
            self.ui.label_28.setText(f'扩充后数据长度：{str(len(self.u4_temp_expan))}')

    def tab4_finished(self):
        if MainWin.ok3:
            if self.ui.checkBox.isChecked():
                MainWin.u4, MainWin.F4 = self.u4_temp_expan, self.F4_temp_expan
                MainWin.ok4 = True
                print('tab4 已扩充')
            else:
                MainWin.u4, MainWin.F4 = self.u4_temp, self.F4_temp
                MainWin.ok4 = True
                print('tab4 不扩充')
            self.tab5_finished()

    def clear_tab4(self):
        MainWin.u4, MainWin.F4 = None, None
        MainWin.ok4 = False
        self.ui.checkBox.setChecked(False)
        self.ui.lineEdit_12.setText('0.2')
        self.ui.label_27.setText('处理前数据长度：')
        self.ui.label_29.setText('单调化后数据长度：')
        self.ui.label_28.setText('扩充后数据长度：不扩充')
        self.MyDel('self.u4_temp', 'self.F4_temp', 'self.u4_temp_expan', 'self.F4_temp_expan')

    # --------------------------------------------------- tab 5 ---------------------------------------------------

    def tab5_finished(self):
        if self.ui.checkBox_2.isChecked():
            self.curve_smooth()
        else:
            self.plot_no_smooth()

    def checkBox2_state_changed(self, state):
        if state == Qt.Checked:
            self.ui.label_30.setEnabled(True)
            self.ui.lineEdit_13.setEnabled(True)
            self.curve_smooth()
        else:
            self.ui.label_30.setEnabled(False)
            self.ui.lineEdit_13.setEnabled(False)
            self.clear_plot_smooth()

    def curve_smooth(self):
        if MainWin.ok4:
            r_window_size = int(self.ui.lineEdit_13.text())
            if r_window_size < 2:
                return 0
            _, self.F5_smooth = smooth(MainWin.u4, MainWin.F4, r_window_size)
            self.plot_smooth()

    def plot_smooth(self):
        if MainWin.ok4:
            self.plot_no_smooth()
            self.pg6.plot(np.arange(0, len(MainWin.u4), 1), self.F5_smooth, pen=self.pen2, name='after smoothing')
            self.pg6.autoRange()

    def plot_no_smooth(self):
        if MainWin.ok4:
            self.pg6.clear()
            self.pg6.addLegend()
            self.pg6.plot(np.arange(0, len(MainWin.u4), 1), MainWin.F4, pen=self.pen1, name='before smoothing')
            self.pg6.autoRange()

    def clear_plot_smooth(self):
        self.MyDel('self.F5_smooth')
        self.plot_no_smooth()

    def tab5_finished(self):
        if MainWin.ok4:
            if self.ui.checkBox_2.isChecked():
                MainWin.u5, MainWin.F5 = MainWin.u4, self.F5_smooth
                MainWin.ok5 = True
                print('tab5 已平滑')
                self.tab6_start()
                return 0
            else:
                MainWin.u5, MainWin.F5 = MainWin.u4, MainWin.F4
                MainWin.ok5 = True
                print('tab5 不平滑')
                self.tab6_start()
                return 0
        else:
            print('tab 5不输出数据')

    def clear_tab5(self):
        MainWin.u5, MainWin.F5 = None, None
        MainWin.ok5 = False
        self.ui.checkBox_2.setChecked(False)
        self.ui.lineEdit_13.setText('5')
        self.pg6.clear()
        self.MyDel('self.F5_smooth')

    # --------------------------------------------------- tab 6 ---------------------------------------------------

    def tab6_start(self):
        self.get_loops_and_energy()

    def get_loops_and_energy(self):
        if MainWin.ok5:
            # 分圈
            if self.ui.checkBox_3.isChecked():
                skip = 0
            else:
                skip = 1
            if self.ui.radioButton_10.isChecked():
                method = 'a'  # 滞回环分圈方法，a-端点为零位移点，b-端点为反向点
            else:
                method = 'b'
            self.u_loops, self.F_loops, self.length_with_skip = loops_division(MainWin.u5, MainWin.F5, skip=skip, method=method)
            # 提取骨架点
            if self.ui.radioButton_3.isChecked():
                skeleton_method = 1
            if self.ui.radioButton_4.isChecked():
                skeleton_method = 2
            tag_gujia = get_skeleton(MainWin.u5[: self.length_with_skip], MainWin.F5[: self.length_with_skip], skeleton_method)
            self.gujia_u, self.gujia_F = MainWin.u5[tag_gujia], MainWin.F5[tag_gujia]
            # 计算耗能
            if self.ui.radioButton_5.isChecked():
                energyMethod = 1
            if self.ui.radioButton_6.isChecked():
                energyMethod = 2
            if self.ui.radioButton_7.isChecked():
                energyMethod = 3
            self.Es, self.Ea, self.zeta = get_dissipated_energy(self.u_loops, self.F_loops, EnergyMethod=energyMethod)
            # 残余变形
            self.residual_pos, self.residual_neg = [], []
            for u_loop, F_loop in zip(self.u_loops, self.F_loops):
                d_pos, d_neg = get_residual_deformation(u_loop, F_loop)
                self.residual_pos.append(d_pos)
                self.residual_neg.append(d_neg)
            # 画图
            self.plot6()

    def plot6(self):
        if MainWin.ok5:
            self.pg7.clear()
            self.pg7.plot(MainWin.u5[: self.length_with_skip], MainWin.F5[: self.length_with_skip], pen=self.pen1)
            self.pg7.plot(self.residual_pos, np.zeros(len(self.residual_pos)), pen=None, symbolBrush='green', symbol='o', symbolSize=10)
            self.pg7.plot(self.residual_neg, np.zeros(len(self.residual_neg)), pen=None, symbolBrush='green', symbol='o', symbolSize=10)
            self.pg7.plot(self.gujia_u, self.gujia_F, pen=None, symbolBrush='red', symbol='o', symbolSize=10)
            self.pg7.autoRange()
            self.pg8.clear()
            self.pg8.plot(np.arange(0, len(self.Es), 1), self.Es, pen=self.pen1, symbol='o', symbolBrush=(68, 114, 196), symbolSize=10)
            self.pg8.autoRange()
            self.pg9.clear()
            self.pg9.plot(np.arange(0, len(self.Ea), 1), self.Ea, pen=self.pen1, symbol='o', symbolBrush=(68, 114, 196), symbolSize=10)
            self.pg9.autoRange()
            self.pg10.clear()
            self.pg10.plot(np.arange(0, len(self.zeta), 1), self.zeta, pen=self.pen1, symbol='o', symbolBrush=(68, 114, 196), symbolSize=10)
            self.pg10.autoRange()

    def tab6_finished(self):
        if MainWin.ok5:
            MainWin.u6, MainWin.F6 = MainWin.u5[: self.length_with_skip], MainWin.F5[: self.length_with_skip]
            MainWin.ok6 = True
            self.tab7_start()

    def clear_tab6(self):
        MainWin.u6, MainWin.F6 = None, None
        MainWin.ok6 = False
        self.ui.radioButton_3.setChecked(True)
        self.ui.checkBox_3.setChecked(True)
        self.ui.radioButton_5.setChecked(True)
        self.pg7.clear()
        self.pg8.clear()
        self.pg9.clear()
        self.pg10.clear()
        self.MyDel('self.u_loops', 'self.F_loops', 'self.length_with_skip', 'self.gujia_u', 'self.gujia_F', 'self.Es', 'self.Ea',\
                   'self.zeta', 'self.residual_pos', 'self.residual_neg')

    # --------------------------------------------------- tab 7 ---------------------------------------------------

    def tab7_start(self):
        self.ui.comboBox.clear()
        self.setup_comboBox()
        self.plot_loop(0)

    def setup_comboBox(self):
        if MainWin.ok6:
            self.ui.comboBox.addItem('完整滞回曲线')
            items = np.arange(1, len(self.u_loops) + 1, 1).tolist()
            items = [f'第{i}圈' for i in items]
            self.ui.comboBox.addItems(items)
            self.current_idx = 0  # 当前下拉列表的索引

    def comboBox_changed(self):
        if MainWin.ok6:
            self.current_idx = self.ui.comboBox.currentIndex()  # 获取下拉列表的索引
            self.current_idx = max(0, self.current_idx)  # self.current_idx会变成-1
            self.plot_loop(self.current_idx)

    def next_loop(self):
        # 下一圈
        if MainWin.ok6:
            self.current_idx += 1
            if self.current_idx > len(self.u_loops):
                self.current_idx = 0
            self.ui.comboBox.setCurrentIndex(self.current_idx)

    def previous_loop(self):
        # 上一圈
        if MainWin.ok6:
            self.current_idx -= 1
            if self.current_idx < 0:
                self.current_idx = len(self.u_loops)
            self.ui.comboBox.setCurrentIndex(self.current_idx)

    def plot_all_loop(self):
        # 绘制滞回曲线
        self.current_idx = 0
        self.ui.comboBox.setCurrentIndex(self.current_idx)

    def plot_loop(self, idx):
        if MainWin.ok6:
            # self.ax12.clear()
            self.pg12.clear()
            if idx == 0:
                current_loop_u, current_loop_F = MainWin.u6, MainWin.F6
                self.pg12.plot(current_loop_u, current_loop_F, pen=self.pen1)
                self.pg12.autoRange()
                self.ui.lineEdit_19.clear()
                self.ui.lineEdit_20.clear()
                self.ui.lineEdit_21.clear()
                self.ui.lineEdit_22.clear()
                self.ui.lineEdit_23.clear()
                self.ui.lineEdit_24.clear()
                self.ui.lineEdit_25.clear()
                self.ui.lineEdit_26.clear()
            else:
                current_loop_u, current_loop_F = self.u_loops[idx - 1], self.F_loops[idx - 1]
                self.pg12.plot(current_loop_u, current_loop_F, pen=self.pen1)
                self.pg12.autoRange()
                self.ui.lineEdit_19.setText(str(round(max(current_loop_u), 5)))
                self.ui.lineEdit_20.setText(str(round(min(current_loop_u), 5)))
                self.ui.lineEdit_21.setText(str(round(max(current_loop_F), 5)))
                self.ui.lineEdit_22.setText(str(round(min(current_loop_F), 5)))
                self.ui.lineEdit_23.setText(str(len(current_loop_u)))
                self.ui.lineEdit_24.setText(str(round(self.Es[idx], 2)))  # 三个耗能指标是从0开始的，所以idx不用减1
                self.ui.lineEdit_25.setText(str(round(self.Ea[idx], 2)))
                self.ui.lineEdit_26.setText(str(round(self.zeta[idx], 5)))

    def tab7_finished(self):
        if MainWin.ok6:
            MainWin.ok7 = True
            MainWin.u7, MainWin.F7 = MainWin.u6, MainWin.F6
            self.tab7_1_start()

    def clear_tab7(self):
        MainWin.u7, MainWin.F7 = None, None
        MainWin.ok7 = False
        self.ui.comboBox.clear()
        self.ui.lineEdit_19.clear()
        self.ui.lineEdit_20.clear()
        self.ui.lineEdit_21.clear()
        self.ui.lineEdit_22.clear()
        self.ui.lineEdit_23.clear()
        self.ui.lineEdit_24.clear()
        self.ui.lineEdit_25.clear()
        self.ui.lineEdit_26.clear()
        self.pg12.clear()
        self.MyDel('self.current_idx')

    # -------------------------------------------------- tab 7_1 --------------------------------------------------

    def tab7_1_start(self):
        pass

    def clicked_add_to_scheme(self):
        if MainWin.ok7:
            self.is_clear = False
            # idx: int or list
            if self.ui.radioButton_8.isChecked():
                # 用常用方案
                idx = self.ui.comboBox_5.currentIndex()
                if not self.scheme:
                    self.scheme = [idx]
                else:
                    self.scheme.append(idx)
                    if len(self.scheme) > 5:
                        self.scheme = self.scheme[:5]
                        return 0
                self.get_skt_deg_result(idx)
                text = self.ui.label_57.text()
                self.ui.label_57.setText(text + '\n%d. 常用方案(%d)' % (len(self.scheme), idx + 1))
                self.ui.comboBox_2.addItem('%d. 常用方案(%d)' % (len(self.scheme), idx + 1))
                self.ui.comboBox_3.addItem('%d. 常用方案(%d)' % (len(self.scheme), idx + 1))
                self.ui.comboBox_4.addItem('%d. 常用方案(%d)' % (len(self.scheme), idx + 1))
                self.data_scheme.append('%d. 常用方案(%d)' % (len(self.scheme), idx + 1))
            else:
                # 用手动方案
                try:
                    idx = self.ui.lineEdit_27.text().split(',')
                    idx = [int(i) for i in idx]
                except:
                    QMessageBox.warning(self, '警告', '输入格式有误！')
                    return 0
                if idx == []:
                    QMessageBox.warning(self, '警告', '输入格式有误！')
                    return 0
                if not self.scheme:
                    self.scheme = [idx]
                else:
                    self.scheme.append(idx)
                text = self.ui.label_57.text()
                n = 0
                for i in self.scheme:
                    if type(i) == list:
                        n += 1  # 计算手动方案的个数
                self.get_skt_deg_result(idx)
                self.ui.label_57.setText(text + '\n%d. 手动方案(%d)' % (len(self.scheme), n))
                self.ui.comboBox_2.addItem('%d. 手动方案(%d)' % (len(self.scheme), n))
                self.ui.comboBox_3.addItem('%d. 手动方案(%d)' % (len(self.scheme), n))
                self.ui.comboBox_4.addItem('%d. 手动方案(%d)' % (len(self.scheme), n))
                self.data_scheme.append('%d. 手动方案(%d)' % (len(self.scheme), n))
            
    def get_skt_deg_result(self, idx):
        if MainWin.ok7:
            if self.is_clear:
                return 0
            else:
                self.is_clear = False
            # 骨架曲线
            if type(idx) == int:
                # 常用方案
                skt_u, skt_F, _ = get_skeleton_curve_1(self.gujia_u, self.gujia_F, idx)
            else:
                # 手动方案
                skt_u, skt_F, _ = get_skeleton_curve_2(self.gujia_u, self.gujia_F, idx)
            self.data_skeleton_curve.append([skt_u, skt_F])
            # 刚度退化
            if type(idx) == int:
                skt_u, skt_F, _ = get_skeleton_curve_1(self.gujia_u, self.gujia_F, idx)
            else:
                skt_u, skt_F, _ = get_skeleton_curve_2(self.gujia_u, self.gujia_F, idx)
            n = int(len(skt_u) / 2)
            stif_deg_x, stif_deg_y = [], []
            for i in range(n):
                stif_deg_x.append(skt_u[n + i])
                deg_val = (skt_F[n + i] - skt_F[n - i - 1]) / (skt_u[n] - skt_u[n - i - 1])  # 割线刚度退化值
                stif_deg_y.append(deg_val)
            self.data_stiffness_degradation.append([stif_deg_x, stif_deg_y])
            # 强度退化
            if type(idx) == int:
                skt_u, skt_F, N = get_skeleton_curve_1(self.gujia_u, self.gujia_F, idx)
            else:
                skt_u, skt_F, N = get_skeleton_curve_2(self.gujia_u, self.gujia_F, idx)
            if N[0] > 0:
                if type(idx) == int:
                    skt_u0, skt_F0, _ = get_skeleton_curve_1(self.gujia_u, self.gujia_F, idx, previous_round=True)
                else:
                    skt_u0, skt_F0, _ = get_skeleton_curve_2(self.gujia_u, self.gujia_F, idx, previous_round=True)
            else:
                # 无上一圈，无强度退化
                self.data_strength_degradation.append([None, None])
                return 0
            strg_deg_x, strg_deg_y = [], []
            if len(skt_u) > len(skt_u0):
                delete_num = int((len(skt_u) - len(skt_u0)) / 2)
                skt_u = skt_u[delete_num: -delete_num]
                skt_F = skt_F[delete_num: -delete_num]
            elif len(skt_u) < len(skt_u0):
                delete_num = int((len(skt_u0) - len(skt_u)) / 2)
                skt_u0 = skt_u0[delete_num: -delete_num]
                skt_F0 = skt_F0[delete_num: -delete_num]
            for i, (F, F0) in enumerate(zip(skt_F, skt_F0)):
                strg_deg_x.append(skt_u[i])
                strg_deg_y.append(F / F0)
            self.data_strength_degradation.append([strg_deg_x, strg_deg_y])

    def clear_all_scheme(self):
        self.is_clear = True
        self.scheme = None
        self.ui.lineEdit_27.clear()
        self.ui.comboBox_2.clear()
        self.ui.comboBox_3.clear()
        self.ui.comboBox_4.clear()
        self.pg13.clear()
        self.pg14.clear()
        self.pg15.clear()
        self.ui.label_57.setText('已选择方案（最多五个）：')
        self.data_scheme = []
        self.data_skeleton_curve = []
        self.data_stiffness_degradation = []
        self.data_strength_degradation = []
        MainWin.ok7_1 = False

    def plot_skeleton(self):
        if MainWin.ok7:
            idx_skt = self.ui.comboBox_2.currentIndex()
            skt_u, skt_F = self.data_skeleton_curve[idx_skt]
            self.pg13.clear()
            self.pg13.plot(MainWin.u7, MainWin.F7, pen=self.pen4, c='g')
            self.pg13.plot(skt_u, skt_F, pen=self.pen5, symbolBrush='red', symbol='o', symbolSize=10)
            # self.pg13.plot([0], [0], pen=None, symbolBrush='red', symbol='o', symbolSize=10)
            self.pg13.autoRange()
            

    def plot_stiffness_degradation(self):
        if MainWin.ok7:
            idx_stif = self.ui.comboBox_3.currentIndex()
            stif_deg_x, stif_deg_y = self.data_stiffness_degradation[idx_stif]
            self.pg14.clear()
            self.pg14.plot(stif_deg_x, stif_deg_y, pen=self.pen1, symbolBrush=(68, 114, 196), symbol='o', symbolSize=10)
            self.pg14.autoRange()
    
    def plot_strength_degradation(self):
        if MainWin.ok7:
            idx_strg = self.ui.comboBox_4.currentIndex()
            strg_deg_x, strg_deg_y = self.data_strength_degradation[idx_strg]
            if not strg_deg_x:
                self.ui.label_56.setText('强度退化：（当前方案无上一圈滞回环，无强度退化）')
                self.pg15.clear()
                return 0
            self.ui.label_56.setText('强度退化：')
            self.pg15.clear()
            self.pg15.plot(strg_deg_x, strg_deg_y, pen=self.pen1, symbolBrush=(68, 114, 196), symbol='o', symbolSize=10)
            self.pg15.autoRange()
            
    def tab7_1_finished(self):
        if MainWin.ok7:
            MainWin.ok7_1 = True
            MainWin.u7_1, MainWin.F7_1 = MainWin.u7, MainWin.F7
            self.tab8_start()

    def clear_tab7_1(self):
        MainWin.u7_1, MainWin.F7_1 = None, None
        MainWin.ok7_1 = False
        self.ui.label_57.setText('已选择方案（最多五个）：')
        self.ui.comboBox_5.setCurrentIndex(0)
        self.ui.radioButton_8.setChecked(True)
        self.ui.lineEdit_27.clear()
        self.ui.comboBox_2.clear()
        self.ui.comboBox_3.clear()
        self.ui.comboBox_4.clear()
        self.pg13.clear()
        self.pg14.clear()
        self.pg15.clear()
        self.scheme = None
        self.is_clear = False
        self.data_skeleton_curve = []
        self.data_stiffness_degradation = []
        self.data_strength_degradation = []
        self.data_scheme = []

    # --------------------------------------------------- tab 8 ---------------------------------------------------

    def tab8_start(self):
        self.result_display()

    def result_display(self):
        if MainWin.ok7_1:
            self.ui.lineEdit_14.setText(str(round(max(MainWin.u7_1), 5)))
            self.ui.lineEdit_15.setText(str(round(min(MainWin.u7_1), 5)))
            self.ui.lineEdit_16.setText(str(round(max(MainWin.F7_1), 5)))
            self.ui.lineEdit_17.setText(str(round(min(MainWin.F7_1), 5)))
            self.ui.lineEdit_18.setText(str(round(len(MainWin.u7_1), 5)))
            self.pg11.clear()
            self.pg11.plot(MainWin.u7_1, MainWin.F7_1, pen=self.pen1)
            self.pg11.autoRange()

    def data_hysteretic_curve(self):
        if MainWin.ok7_1:
            data = np.zeros((len(MainWin.u7_1), 2))
            data[:, 0], data[:, 1] = MainWin.u7_1, MainWin.F7_1
            self.WinData = WinData(data, content='滞回曲线')
            _translate = QCoreApplication.translate
            self.WinData.setWindowTitle(_translate("win_getData", "滞回曲线"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(0).setText(_translate("win_getData", "位移"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(1).setText(_translate("win_getData", "力"))
            self.WinData.ui_data.tableWidget.setRowCount(len(MainWin.u7_1))
            for i in range(len(MainWin.u7_1)):
                item_x = QTableWidgetItem(str(round(MainWin.u7_1[i], 5)))
                item_y = QTableWidgetItem(str(round(MainWin.F7_1[i], 5)))
                self.WinData.ui_data.tableWidget.setItem(i, 0, item_x)
                self.WinData.ui_data.tableWidget.setItem(i, 1, item_y)
            self.WinData.exec_()
            del self.WinData
        else:
            QMessageBox.warning(self, '警告', '没有数据！')

    def data_skeleton_points(self):
        if MainWin.ok7_1:
            data = np.zeros((len(self.gujia_u) + 1, 2))
            data[:, 0], data[:, 1] = np.insert(self.gujia_u, 0, 0), np.insert(self.gujia_F, 0, 0)
            self.WinData = WinData(data, content='骨架点')
            _translate = QCoreApplication.translate
            self.WinData.setWindowTitle(_translate("win_getData", "骨架点"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(0).setText(_translate("win_getData", "位移"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(1).setText(_translate("win_getData", "力"))
            self.WinData.ui_data.tableWidget.setRowCount(len(self.gujia_u) + 1)
            for i in range(len(data)):
                item_x = QTableWidgetItem(str(round(data[i, 0], 5)))
                item_y = QTableWidgetItem(str(round(data[i, 1], 5)))
                self.WinData.ui_data.tableWidget.setItem(i, 0, item_x)
                self.WinData.ui_data.tableWidget.setItem(i, 1, item_y)
            self.WinData.exec_()
            del self.WinData
        else:
            QMessageBox.warning(self, '警告', '没有数据！')

    def data_Es(self):
        if MainWin.ok7_1:
            data = np.zeros((len(self.Es), 2))
            data[:, 0], data[:, 1] = np.arange(0, len(self.Es), 1), self.Es
            self.WinData = WinData(data, content='单圈耗能')
            _translate = QCoreApplication.translate
            self.WinData.setWindowTitle(_translate("win_getData", "单圈耗能"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(0).setText(_translate("win_getData", "循环周数"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(1).setText(_translate("win_getData", "单圈耗能"))
            self.WinData.ui_data.tableWidget.setRowCount(len(self.Es))
            for i in range(len(self.Es)):
                item_x = QTableWidgetItem(str(i))
                item_y = QTableWidgetItem(str(round(self.Es[i], 5)))
                self.WinData.ui_data.tableWidget.setItem(i, 0, item_x)
                self.WinData.ui_data.tableWidget.setItem(i, 1, item_y)
            self.WinData.exec_()
            del self.WinData
        else:
            QMessageBox.warning(self, '警告', '没有数据！')

    def data_Ea(self):
        if MainWin.ok7_1:
            data = np.zeros((len(self.Ea), 2))
            data[:, 0], data[:, 1] = np.arange(0, len(self.Ea), 1), self.Ea
            self.WinData = WinData(data, content='累积耗能')
            _translate = QCoreApplication.translate
            self.WinData.setWindowTitle(_translate("win_getData", "累积耗能"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(0).setText(_translate("win_getData", "循环周数"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(1).setText(_translate("win_getData", "累积耗能"))
            self.WinData.ui_data.tableWidget.setRowCount(len(self.Ea))
            for i in range(len(self.Ea)):
                item_x = QTableWidgetItem(str(i))
                item_y = QTableWidgetItem(str(round(self.Ea[i], 5)))
                self.WinData.ui_data.tableWidget.setItem(i, 0, item_x)
                self.WinData.ui_data.tableWidget.setItem(i, 1, item_y)
            self.WinData.exec_()
            del self.WinData
        else:
            QMessageBox.warning(self, '警告', '没有数据！')

    def data_zeta(self):
        if MainWin.ok7_1:
            data = np.zeros((len(self.zeta), 2))
            data[:, 0], data[:, 1] = np.arange(0, len(self.zeta), 1), self.zeta
            self.WinData = WinData(data, content='等效粘滞阻尼系数')
            _translate = QCoreApplication.translate
            self.WinData.setWindowTitle(_translate("win_getData", "等效阻尼"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(0).setText(_translate("win_getData", "循环周数"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(1).setText(_translate("win_getData", "等效阻尼"))
            self.WinData.ui_data.tableWidget.setRowCount(len(self.zeta))
            for i in range(len(self.zeta)):
                item_x = QTableWidgetItem(str(i))
                item_y = QTableWidgetItem(str(round(self.zeta[i], 5)))
                self.WinData.ui_data.tableWidget.setItem(i, 0, item_x)
                self.WinData.ui_data.tableWidget.setItem(i, 1, item_y)
            self.WinData.exec_()
            del self.WinData
        else:
            QMessageBox.warning(self, '警告', '没有数据！')

    def data_residual_deformation(self):
        if MainWin.ok7_1:
            data = np.zeros((len(self.residual_pos), 3))
            data[:, 0], data[:, 1], data[:, 2] = np.arange(1, len(self.residual_pos) + 1, 1), self.residual_pos, self.residual_neg
            self.WinData = WinData(data, content='残余变形', isResidual=True)
            _translate = QCoreApplication.translate
            self.WinData.setWindowTitle(_translate("win_getData", "残余变形"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(0).setText(_translate("win_getData", "循环周数"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(1).setText(_translate("win_getData", "正向残余变形"))
            self.WinData.ui_data.tableWidget.horizontalHeaderItem(2).setText(_translate("win_getData", "负向残余变形"))
            self.WinData.ui_data.tableWidget.setRowCount(len(self.residual_pos))
            for i in range(len(self.residual_pos)):
                item_1 = QTableWidgetItem(str(i + 1))
                item_2 = QTableWidgetItem(str(round(self.residual_pos[i], 5)))
                item_3 = QTableWidgetItem(str(round(self.residual_neg[i], 5)))
                self.WinData.ui_data.tableWidget.setItem(i, 0, item_1)
                self.WinData.ui_data.tableWidget.setItem(i, 1, item_2)
                self.WinData.ui_data.tableWidget.setItem(i, 2, item_3)
            self.WinData.exec_()
            del self.WinData
        else:
            QMessageBox.warning(self, '警告', '没有数据！')

    def export_all_data(self):
        if MainWin.ok7_1:
            self.ui.pushButton_24.setText('正在导出...')
            self.ui.pushButton_24.setEnabled(False)
            output_file = QFileDialog.getExistingDirectory(self, '选择保存路径文件夹')
            if not output_file:
                return 0
            np.savetxt(f'{output_file}/滞回曲线.txt', np.column_stack((MainWin.u7_1, MainWin.F7_1)))
            np.savetxt(f'{output_file}/骨架点.txt', np.column_stack((self.gujia_u, self.gujia_F)))
            np.savetxt(f'{output_file}/单圈耗能.txt', self.Es)
            np.savetxt(f'{output_file}/累积耗能.txt', self.Ea)
            np.savetxt(f'{output_file}/等效粘滞阻尼系数.txt', self.zeta)
            np.savetxt(f'{output_file}/残余变形.txt', np.column_stack((np.arange(0, len(self.residual_pos), 1), self.residual_pos, self.residual_neg)))
            if not os.path.exists(f'{output_file}/各圈滞回环'):
                os.mkdir(f'{output_file}/各圈滞回环')
            for i, (u_loop, F_loop) in enumerate(zip(self.u_loops, self.F_loops)):
                np.savetxt(f'{output_file}/各圈滞回环/第{i+1}圈滞回环.txt', np.column_stack((u_loop, F_loop)))
            export_result = self.creat_excel(output_file)
            if export_result:
                QMessageBox.information(self, '提示', '已导出所有数据！')
        else:
            QMessageBox.warning(self, '警告', '没有数据！')

    def creat_excel(self, output_file):
        if MainWin.ok7_1:
            wb = px.Workbook()
            # 1 滞回曲线
            ws1 = wb.active
            ws1.title = '滞回曲线'
            ws1['A1'] = 'u'
            ws1['B1'] = 'F'
            self.write_col_to_excel(ws1, MainWin.u7_1, 2, 1)
            self.write_col_to_excel(ws1, MainWin.F7_1, 2, 2)
            for row in ws1.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            # 2 骨架曲线&退化曲线
            ws2 = wb.create_sheet('骨架曲线&退化曲线', 1)
            for i, title in enumerate(self.data_scheme):
                title = self.data_scheme[i][3:]
                ws2.cell(row=1, column=1+7*i, value=title)
                ws2.cell(row=3, column=1+7*i, value='骨架曲线')
                ws2.cell(row=3, column=3+7*i, value='刚度退化曲线')
                ws2.cell(row=3, column=5+7*i, value='强度退化曲线')
                ws2.cell(row=4, column=1+7*i, value='u')
                ws2.cell(row=4, column=2+7*i, value='F')
                ws2.cell(row=4, column=3+7*i, value='u')
                ws2.cell(row=4, column=4+7*i, value='Ke')
                ws2.cell(row=4, column=5+7*i, value='u')
                ws2.cell(row=4, column=6+7*i, value='lamda')
                ws2.merge_cells(start_row=1, start_column=1+7*i, end_row=2, end_column=6+7*i)
                ws2.merge_cells(start_row=3, start_column=1+7*i, end_row=3, end_column=2+7*i)
                ws2.merge_cells(start_row=3, start_column=3+7*i, end_row=3, end_column=4+7*i)
                ws2.merge_cells(start_row=3, start_column=5+7*i, end_row=3, end_column=6+7*i)
                self.write_col_to_excel(ws2, self.data_skeleton_curve[i][0], 5, 1+7*i)
                self.write_col_to_excel(ws2, self.data_skeleton_curve[i][1], 5, 2+7*i)
                self.write_col_to_excel(ws2, self.data_stiffness_degradation[i][0], 5, 3+7*i)
                self.write_col_to_excel(ws2, self.data_stiffness_degradation[i][1], 5, 4+7*i)
                if self.data_strength_degradation[i][0] is None:
                    ws2.cell(row=5, column=5+7*i, value='当前方案无强度退化')
                    ws2.merge_cells(start_row=5, start_column=5+7*i, end_row=5, end_column=6+7*i)
                else:
                    self.write_col_to_excel(ws2, self.data_strength_degradation[i][0], 5, 5+7*i)
                    self.write_col_to_excel(ws2, self.data_strength_degradation[i][1], 5, 6+7*i)
            for row in ws2.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            # 3 耗能指标&残余变形
            ws3 = wb.create_sheet('耗能指标&其他', 2)
            ws3.cell(row=1, column=1, value='单圈耗能')
            ws3.cell(row=1, column=3, value='累积耗能')
            ws3.cell(row=1, column=5, value='等效粘滞阻尼系数')
            ws3.cell(row=1, column=7, value='所有骨架点')
            ws3.cell(row=1, column=9, value='残余变形')
            ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws3.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
            ws3.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
            ws3.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
            ws3.merge_cells(start_row=1, start_column=9, end_row=1, end_column=11)
            ws3.cell(row=2, column=1, value='cycle')
            ws3.cell(row=2, column=2, value='Es')
            ws3.cell(row=2, column=3, value='cycle')
            ws3.cell(row=2, column=4, value='Ea')
            ws3.cell(row=2, column=5, value='cycle')
            ws3.cell(row=2, column=6, value='zeta')
            ws3.cell(row=2, column=7, value='u')
            ws3.cell(row=2, column=8, value='F')
            ws3.cell(row=2, column=9, value='cycle')
            ws3.cell(row=2, column=10, value='+')
            ws3.cell(row=2, column=11, value='-')
            self.write_col_to_excel(ws3, np.arange(0, len(self.Es), 1), 3, 1)
            self.write_col_to_excel(ws3, self.Es, 3, 2)
            self.write_col_to_excel(ws3, np.arange(0, len(self.Ea), 1), 3, 3)
            self.write_col_to_excel(ws3, self.Ea, 3, 4)
            self.write_col_to_excel(ws3, np.arange(0, len(self.zeta), 1), 3, 5)
            self.write_col_to_excel(ws3, self.zeta, 3, 6)
            self.write_col_to_excel(ws3, self.gujia_u, 3, 7)
            self.write_col_to_excel(ws3, self.gujia_F, 3, 8)
            data = np.zeros((len(self.residual_pos), 3))
            data[:, 0], data[:, 1], data[:, 2] = np.arange(1, len(self.residual_pos) + 1, 1), self.residual_pos, self.residual_neg
            self.write_col_to_excel(ws3, np.array(data[:, 0]), 3, 9)
            self.write_col_to_excel(ws3, np.array(data[:, 1]), 3, 10)
            self.write_col_to_excel(ws3, np.array(data[:, 2]), 3, 11)
            for row in ws3.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            # 4 各圈滞回环
            ws4 = wb.create_sheet('各圈滞回环', 3)
            for i, (u_loop, F_loop) in enumerate(zip(self.u_loops, self.F_loops)):
                ws4.cell(row=1, column=1+i*2, value=f'第{i+1}圈')
                ws4.cell(row=2, column=1+i*2, value='u')
                ws4.cell(row=2, column=2+i*2, value='F')
                ws4.merge_cells(start_row=1, start_column=1+i*2, end_row=1, end_column=2+i*2)
                self.write_col_to_excel(ws4, u_loop, 3, 1+i*2)
                self.write_col_to_excel(ws4, F_loop, 3, 2+i*2)
            for row in ws4.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            # 导出
            try:
                wb.save(f'{output_file}/数据汇总.xlsx')
                self.ui.pushButton_24.setText('导出所有数据')
                self.ui.pushButton_24.setEnabled(True)
                return True
            except:
                QMessageBox.warning(self, '警告', '无法导出excel，若正在打开excel请关闭。')
                self.ui.pushButton_24.setText('导出所有数据')
                self.ui.pushButton_24.setEnabled(True)
                return False

    def write_col_to_excel(self, ws, data, row, col):
        # 写入一列数据到excel
        for i, val in enumerate(data):
            ws.cell(row=row+i, column=col, value=val)

    def clear_tab8(self):
        self.ui.lineEdit_14.clear()
        self.ui.lineEdit_15.clear()
        self.ui.lineEdit_16.clear()
        self.ui.lineEdit_17.clear()
        self.ui.lineEdit_18.clear()
        self.pg11.clear()    

    # --------------------------------------------------- Menu ---------------------------------------------------

    def show_WinAbout(self):
        self.WinAbout = WinAbout()
        self.WinAbout.exec_()

    def show_WinHelp(self):
        self.WinHelp = WinHelp()
        self.WinHelp.exec_()
        
    # ------------------------------------------------- WinData ---------------------------------------------------

class WinData(QDialog):
    def __init__(self, data, content, isResidual=False):
        super().__init__()
        self.data, self.content = data, content
        self.ui_data = Ui_WinData()
        self.ui_data.setupUi(self)
        header = self.ui_data.tableWidget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        if isResidual:
            self.resize(600, 500)
            self.ui_data.tableWidget.insertColumn(2)  # 如果统计残余变形，则增加一列
            item = QTableWidgetItem()
            self.ui_data.tableWidget.setHorizontalHeaderItem(2, item)
        self.ui_data.tableWidget.setEditTriggers(self.ui_data.tableWidget.NoEditTriggers)  # 禁止编辑
        self.ui_data.pushButton.clicked.connect(self.data_export)
        self.ui_data.pushButton_2.clicked.connect(self.data_copy)
        self.ui_data.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ui_data.tableWidget.customContextMenuRequested.connect(self.showContextMenu)

    def showContextMenu(self, pos):
        # 定义上下文菜单
        context_menu = QMenu(self)
        copy_action = context_menu.addAction("复制")
        menu_size = context_menu.sizeHint()  # 获取菜单的大小
        global_pos = self.ui_data.tableWidget.mapToGlobal(pos)  # 获取全局坐标
        adjusted_pos = global_pos + QPoint(menu_size.width() // 2, menu_size.height() // 2)  # 调整位置以使其出现在鼠标指针的右下方
        action = context_menu.exec_(adjusted_pos)
        if action == copy_action:
            self.copy_to_clipboard()

    def copy_to_clipboard(self):
        # 幅值至剪切板
        selected = self.ui_data.tableWidget.selectedItems()
        if selected:
            row_start = selected[0].row()
            row_end = selected[-1].row()
            col_start = selected[0].column()
            col_end = selected[-1].column()
            clipboard_text = ""
            for r in range(row_start, row_end + 1):
                row_data = []
                for c in range(col_start, col_end + 1):
                    item = self.ui_data.tableWidget.item(r, c)
                    if item and item.text():
                        row_data.append(item.text())
                clipboard_text += "\t".join(row_data) + "\n"
            QApplication.clipboard().setText(clipboard_text)

    def keyPressEvent(self, event):
        # ctrl+C复制
        if event.key() == Qt.Key_C and event.modifiers() == Qt.ControlModifier:
            self.copy_to_clipboard()
        else:
            super().keyPressEvent(event)

    def data_export(self):
        if MainWin.ok6:
            fileName, _ = QFileDialog.getSaveFileName(self, "保存", self.content, "文本文档 (*.txt)")
            if fileName:
                np.savetxt(fileName, self.data)
                QMessageBox.information(self, '提示', f'数据已保存至{fileName}')
        else:
            QMessageBox.warning(self, '警告', '无数据！')

    def data_copy(self):
        if MainWin.ok6:
            QApplication.clipboard().setText("\n".join(["\t".join(map(str, row)) for row in self.data]))
            QMessageBox.information(self, '提示', '已复制。')
        else:
            QMessageBox.warning(self, '警告', '无数据！')

# ------------------------------------------------ WinAbout -------------------------------------------------

class WinAbout(QDialog):
    def __init__(self):
        super().__init__()
        self.ui_WinAbout = Ui_WinAbout()
        self.ui_WinAbout.setupUi(self)
        self.init_ui_about()

    def init_ui_about(self):
        # 替换日期及版本
        text = self.ui_WinAbout.label_3.text()
        text = text.replace('Version', MainWin.Version)
        text = text.replace('data', MainWin.date)
        self.ui_WinAbout.label_3.setText(text)

# ----------------------------------------------- WinHelp ----------------------------------------------------

class WinHelp(QDialog):
    def __init__(self):
        super().__init__()
        self.ui_WinHelp = Ui_WinHelp()
        self.ui_WinHelp.setupUi(self)

